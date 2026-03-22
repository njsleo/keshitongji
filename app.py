import streamlit as st
import pandas as pd
import io
import re
import datetime
import zipfile
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# ================= 1. 网页基础设置 & 究极 UI 美化 =================
st.set_page_config(page_title="教师课时管理系统", page_icon="🎓", layout="wide")

st.markdown("""
<style>
    [data-testid="stHeader"] { background-color: transparent !important; }
    .stApp { background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%); font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; }
    [data-testid="block-container"] { padding-top: 0rem !important; padding-left: 2rem !important; padding-right: 2rem !important; max-width: 98% !important; }
    .custom-header {
        background: linear-gradient(90deg, #1e3c72 0%, #2a5298 100%); color: white; padding: 1.5rem; border-radius: 0 0 15px 15px; 
        text-align: center; box-shadow: 0 6px 20px rgba(30, 60, 114, 0.2); margin-bottom: 25px; margin-top: -15px; 
        font-size: 28px; font-weight: 900; letter-spacing: 2px;
    }
    [data-testid="stSidebar"] { background-color: rgba(255, 255, 255, 0.9) !important; border-right: 1px solid rgba(255,255,255,0.5); box-shadow: 2px 0 15px rgba(0,0,0,0.05); }
    div.stButton > button {
        white-space: nowrap !important; font-size: 13px !important; padding: 4px 12px !important; min-height: 32px !important; 
        height: 32px !important; width: 100% !important; background-color: rgba(255, 255, 255, 0.95) !important;      
        color: #334155 !important; border: 1px solid #cbd5e1 !important; border-radius: 20px !important; box-shadow: 0 2px 4px rgba(0,0,0,0.02) !important;
        transition: all 0.3s ease !important; 
    }
    div.stButton > button:hover {
        background: linear-gradient(90deg, #2a5298 0%, #1e3c72 100%) !important; color: white !important; border-color: #1e3c72 !important;
        transform: translateY(-2px); box-shadow: 0 6px 12px rgba(42, 82, 152, 0.25) !important;
    }
    div[data-testid="stDownloadButton"] > button {
        background: linear-gradient(135deg, #f6d365 0%, #fda085 100%) !important; color: white !important; border: none !important; letter-spacing: 1px;
        border-radius: 8px !important; box-shadow: 0 4px 15px rgba(253, 160, 133, 0.4) !important; font-size: 14px !important;
    }
    div[data-testid="stDownloadButton"] > button:hover { background: linear-gradient(135deg, #fda085 0%, #f6d365 100%) !important; transform: scale(1.02); }
    .row-title { font-size: 14px; font-weight: bold; color: #1e293b; text-align: right; padding-top: 6px; padding-right: 12px; white-space: nowrap; }
    [data-testid="stDataFrame"] { background: white; padding: 10px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.03); border: 1px solid #f1f5f9; }
    [data-testid="column"] { padding: 0 5px !important; }
    .sidebar-subtitle { font-size: 15px; font-weight: bold; color: #2a5298; margin-bottom: 5px; margin-top: 10px; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="custom-header">🎓 教师排课智能读取与精准统计系统</div>', unsafe_allow_html=True)

if 'all_sheets' not in st.session_state: st.session_state['all_sheets'] = None
if 'current_sheet' not in st.session_state: st.session_state['current_sheet'] = None
if 'global_mode' not in st.session_state: st.session_state['global_mode'] = False
if 'teacher_mode' not in st.session_state: st.session_state['teacher_mode'] = False
if 'batch_export_mode' not in st.session_state: st.session_state['batch_export_mode'] = False
if 'search_teacher' not in st.session_state: st.session_state['search_teacher'] = ""

# ================= 辅助函数 =================
def col2num(col_str):
    expn = 0; col_num = 0
    for char in reversed(str(col_str).upper().strip()):
        if 'A' <= char <= 'Z':
            col_num += (ord(char) - ord('A') + 1) * (26 ** expn)
            expn += 1
    return col_num - 1 if col_num > 0 else 0

def filter_first_week(schedule_list):
    if not schedule_list: return schedule_list
    valid_dates = [x['日期'] for x in schedule_list if x['日期']]
    if not valid_dates: return schedule_list
    min_date = min(valid_dates)
    start_of_week = min_date - datetime.timedelta(days=min_date.weekday())
    end_of_week = start_of_week + datetime.timedelta(days=6)
    return [x for x in schedule_list if x['日期'] and (start_of_week <= x['日期'] <= end_of_week)]

# -------- Excel 渲染模块 --------
def convert_df_to_excel_pro(df, sheet_name, title):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        write_index = True if df.index.name or (isinstance(df.index, pd.MultiIndex)) else False
        df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=write_index)
        worksheet = writer.sheets[sheet_name]
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        num_index_cols = len(df.index.names) if write_index else 0
        max_col = len(df.columns) + num_index_cols
        max_row = len(df) + 3 
        cell = worksheet.cell(row=1, column=1, value=title)
        cell.font = Font(size=18, bold=True, color="000000")
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        cell.alignment = center_align
        worksheet.row_dimensions[1].height = 40 
        worksheet.row_dimensions[3].height = 30
        for col_idx in range(1, max_col + 1):
            c = worksheet.cell(row=3, column=col_idx)
            c.fill = header_fill; c.font = header_font; c.alignment = center_align; c.border = thin_border
        for r_idx in range(4, max_row + 1):
            worksheet.row_dimensions[r_idx].height = 45 
            for c_idx in range(1, max_col + 1):
                c = worksheet.cell(row=r_idx, column=c_idx)
                val = str(c.value).strip()
                if val.lower() in ['nan', 'none', 'nat', '0', '0.0']: c.value = ""
                c.alignment = center_align; c.border = thin_border
                if c_idx <= num_index_cols: c.font = Font(bold=True) 
        for i in range(1, max_col + 1):
            worksheet.column_dimensions[get_column_letter(i)].width = 14 if i <= num_index_cols else 24 
    return output.getvalue()

def convert_multiple_dfs_to_excel_pro(df_dict):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, (df, title) in df_dict.items():
            safe_sheet_name = re.sub(r'[\\/*?\[\]:]', '', str(sheet_name))[:31]
            if not safe_sheet_name: safe_sheet_name = "未命名教师"
            write_index = True if df.index.name or (isinstance(df.index, pd.MultiIndex)) else False
            df.to_excel(writer, sheet_name=safe_sheet_name, startrow=2, index=write_index)
            worksheet = writer.sheets[safe_sheet_name]
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            header_fill = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            num_index_cols = len(df.index.names) if write_index else 0
            max_col = len(df.columns) + num_index_cols
            max_row = len(df) + 3 
            cell = worksheet.cell(row=1, column=1, value=title)
            cell.font = Font(size=18, bold=True, color="000000")
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            cell.alignment = center_align
            worksheet.row_dimensions[1].height = 40 
            worksheet.row_dimensions[3].height = 30
            for col_idx in range(1, max_col + 1):
                c = worksheet.cell(row=3, column=col_idx)
                c.fill = header_fill; c.font = header_font; c.alignment = center_align; c.border = thin_border
            for r_idx in range(4, max_row + 1):
                worksheet.row_dimensions[r_idx].height = 45 
                for c_idx in range(1, max_col + 1):
                    c = worksheet.cell(row=r_idx, column=c_idx)
                    val = str(c.value).strip()
                    if val.lower() in ['nan', 'none', 'nat', '0', '0.0']: c.value = ""
                    c.alignment = center_align; c.border = thin_border
                    if c_idx <= num_index_cols: c.font = Font(bold=True) 
            for i in range(1, max_col + 1):
                worksheet.column_dimensions[get_column_letter(i)].width = 14 if i <= num_index_cols else 24 
    return output.getvalue()

def convert_stacked_dfs_to_excel_pro(df_list, sheet_name="全校总表 (垂直合并)"):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        current_row = 0
        for teacher_name, df, title in df_list:
            write_index = True if df.index.name or (isinstance(df.index, pd.MultiIndex)) else False
            df.to_excel(writer, sheet_name=sheet_name, startrow=current_row + 2, index=write_index)
            worksheet = writer.sheets[sheet_name]
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            header_fill = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            num_index_cols = len(df.index.names) if write_index else 0
            max_col = len(df.columns) + num_index_cols
            max_row_for_df = len(df) + 3 
            cell = worksheet.cell(row=current_row + 1, column=1, value=title)
            cell.font = Font(size=18, bold=True, color="000000")
            worksheet.merge_cells(start_row=current_row + 1, start_column=1, end_row=current_row + 1, end_column=max_col)
            cell.alignment = center_align
            worksheet.row_dimensions[current_row + 1].height = 40 
            worksheet.row_dimensions[current_row + 3].height = 30
            for col_idx in range(1, max_col + 1):
                c = worksheet.cell(row=current_row + 3, column=col_idx)
                c.fill = header_fill; c.font = header_font; c.alignment = center_align; c.border = thin_border
            for r_idx in range(current_row + 4, current_row + max_row_for_df + 1):
                worksheet.row_dimensions[r_idx].height = 45 
                for c_idx in range(1, max_col + 1):
                    c = worksheet.cell(row=r_idx, column=c_idx)
                    val = str(c.value).strip()
                    if val.lower() in ['nan', 'none', 'nat', '0', '0.0']: c.value = ""
                    c.alignment = center_align; c.border = thin_border
                    if c_idx <= num_index_cols: c.font = Font(bold=True) 
            for i in range(1, max_col + 1):
                worksheet.column_dimensions[get_column_letter(i)].width = 14 if i <= num_index_cols else 24 
            current_row += max_row_for_df + 3 
    return output.getvalue()

# 【神级核对表清洗排版】：自动计算变动课时数，并加密锁定保护！
def render_verification_sheet(worksheet, grid_df, class_name, f_start, f_end):
    R, C = grid_df.shape
    title = f"【{class_name}】课时核对表 ({f_start} 至 {f_end})"
    
    title_font = Font(size=18, bold=True, color="000000")
    header_font = Font(color="FFFFFF", bold=True, size=13)
    std_fill = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid") 
    act_fill = PatternFill(start_color="D35400", end_color="D35400", fill_type="solid") 
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    unlocked = Protection(locked=False)
    
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=C)
    cell = worksheet.cell(row=1, column=1, value=title)
    cell.font = title_font; cell.alignment = center_align
    worksheet.row_dimensions[1].height = 40
    
    std_start_row = 3
    worksheet.merge_cells(start_row=std_start_row, start_column=1, end_row=std_start_row, end_column=C)
    cell = worksheet.cell(row=std_start_row, column=1, value="⬇️ === 标准课表 (此区域已锁定保护，禁止修改) === ⬇️")
    cell.font = header_font; cell.fill = std_fill; cell.alignment = center_align
    worksheet.row_dimensions[std_start_row].height = 35
    
    for r in range(R):
        worksheet.row_dimensions[std_start_row + 1 + r].height = 25 if r > 1 else 30
        for c in range(C):
            val = str(grid_df.iloc[r, c]).strip()
            if val.lower() in ['nan', 'none', 'nat', '0', '0.0']: val = ""
            cell = worksheet.cell(row=std_start_row + 1 + r, column=c + 1, value=val)
            cell.alignment = center_align; cell.border = thin_border
            if c < 2 or r == 0: cell.font = Font(bold=True)
            elif r == 1 and "星期" in val: cell.font = Font(bold=True)
    
    act_start_row = std_start_row + 1 + R + 2
    worksheet.merge_cells(start_row=act_start_row, start_column=1, end_row=act_start_row, end_column=C)
    cell = worksheet.cell(row=act_start_row, column=1, value="✍️ === 实际上课核对 (仅允许修改下方空白课表区，修改后自动红底黑字标红) === ✍️")
    cell.font = header_font; cell.fill = act_fill; cell.alignment = center_align
    worksheet.row_dimensions[act_start_row].height = 35
    
    for r in range(R):
        worksheet.row_dimensions[act_start_row + 1 + r].height = 25 if r > 1 else 30
        for c in range(C):
            val = str(grid_df.iloc[r, c]).strip()
            if val.lower() in ['nan', 'none', 'nat', '0', '0.0']: val = ""
            cell = worksheet.cell(row=act_start_row + 1 + r, column=c + 1, value=val)
            cell.alignment = center_align; cell.border = thin_border
            if c < 2 or r == 0: cell.font = Font(bold=True)
            elif r == 1 and "星期" in val: cell.font = Font(bold=True)
                
            if r >= 2 and c >= 2:
                cell.protection = unlocked
    
    for i in range(1, C + 1):
        worksheet.column_dimensions[get_column_letter(i)].width = 14 if i <= 2 else 20
    
    # 【黑字红底的极强对比度警示】
    start_col_letter = get_column_letter(3)
    end_col_letter = get_column_letter(C)
    actual_data_start_row = act_start_row + 3 
    actual_data_end_row = act_start_row + R
    std_data_start_row = std_start_row + 3
    std_data_end_row = std_start_row + R
    
    range_string = f"{start_col_letter}{actual_data_start_row}:{end_col_letter}{actual_data_end_row}"
    formula = f'{start_col_letter}{actual_data_start_row}<>{start_col_letter}{std_data_start_row}'
    
    pure_red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    bold_black_font = Font(color="000000", bold=True)
    
    rule = FormulaRule(formula=[formula], stopIfTrue=True, fill=pure_red_fill, font=bold_black_font)
    worksheet.conditional_formatting.add(range_string, rule)
    
    # ================= 【本次更新：自动统计标红异常课时数】 =================
    summary_row_1 = act_start_row + R + 1
    summary_row_2 = act_start_row + R + 2
    
    summary_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # 浅黄高亮底色
    
    # 第一行：每日统计
    worksheet.merge_cells(start_row=summary_row_1, start_column=1, end_row=summary_row_1, end_column=2)
    cell = worksheet.cell(row=summary_row_1, column=1, value="🚨 每日异常标红数：")
    cell.font = Font(bold=True, color="000000")
    cell.alignment = Alignment(horizontal='right', vertical='center')
    
    for c in range(3, C + 1):
        col_letter = get_column_letter(c)
        # 用 SUMPRODUCT 数组公式自动统计每天被改动的格子数量
        sum_formula = f'=SUMPRODUCT(--({col_letter}{actual_data_start_row}:{col_letter}{actual_data_end_row}<>{col_letter}{std_data_start_row}:{col_letter}{std_data_end_row}))'
        cell = worksheet.cell(row=summary_row_1, column=c, value=sum_formula)
        cell.font = Font(bold=True, color="FF0000", size=12)
        cell.alignment = center_align

    # 第二行：全周总计
    worksheet.merge_cells(start_row=summary_row_2, start_column=1, end_row=summary_row_2, end_column=2)
    cell = worksheet.cell(row=summary_row_2, column=1, value="📢 本班本周异常总计：")
    cell.font = Font(bold=True, color="000000")
    cell.alignment = Alignment(horizontal='right', vertical='center')
    
    worksheet.merge_cells(start_row=summary_row_2, start_column=3, end_row=summary_row_2, end_column=C)
    total_formula = f'=SUM({start_col_letter}{summary_row_1}:{end_col_letter}{summary_row_1})'
    cell = worksheet.cell(row=summary_row_2, column=3, value=total_formula)
    cell.font = Font(bold=True, color="FF0000", size=14)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.number_format = '0 "节"' # 直接显示为 X 节
    
    # 统一画边框、刷底色
    for r_idx in [summary_row_1, summary_row_2]:
        worksheet.row_dimensions[r_idx].height = 35
        for c_idx in range(1, C + 1):
            c_cell = worksheet.cell(row=r_idx, column=c_idx)
            c_cell.fill = summary_fill
            c_cell.border = thin_border
            
    # 【开启全局锁定机制】密码设为 123，保护公式不被破坏
    worksheet.protection.sheet = True
    worksheet.protection.password = '123'

def convert_verification_dfs_to_excel(class_grids, f_start, f_end):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for class_name, grid_df in class_grids.items():
            safe_sheet_name = re.sub(r'[\\/*?\[\]:]', '', str(class_name))[:31]
            if not safe_sheet_name: safe_sheet_name = "班级核对表"
            pd.DataFrame().to_excel(writer, sheet_name=safe_sheet_name)
            worksheet = writer.sheets[safe_sheet_name]
            render_verification_sheet(worksheet, grid_df, class_name, f_start, f_end)
    return output.getvalue()

def convert_verification_to_zip(class_grids, f_start, f_end):
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for class_name, grid_df in class_grids.items():
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                pd.DataFrame().to_excel(writer, sheet_name="核对表")
                worksheet = writer.sheets["核对表"]
                render_verification_sheet(worksheet, grid_df, class_name, f_start, f_end)
            excel_data = excel_buffer.getvalue()
            file_name = f"【{class_name}】课时核对表_{f_start}至{f_end}.xlsx"
            zip_file.writestr(file_name, excel_data)
    return zip_buffer.getvalue()

# ================= 智能识别与清洗引擎 =================
def clean_excel_data(df):
    is_schedule = False
    for i in range(min(20, len(df))): 
        row_str = " ".join(str(x) for x in df.iloc[i].values)
        if "星期" in row_str or re.search(r'(\d{4})[-/年\.](\d{1,2})[-/月\.](\d{1,2})', row_str):
            is_schedule = True; break
            
    if is_schedule:
        new_cols = [get_column_letter(i+1) for i in range(len(df.columns))]
        df.columns = new_cols
        return df
    else:
        header_idx = -1
        for i in range(min(10, len(df))):
            if any(k in str(df.iloc[i].values) for k in ["姓名", "科目", "类别", "课数"]):
                header_idx = i; break
        if header_idx != -1:
            raw_cols = df.iloc[header_idx].tolist()
            df = df.iloc[header_idx + 1:].reset_index(drop=True)
        else:
            raw_cols = df.iloc[0].tolist() if len(df)>0 else []
            
        new_cols = []
        for idx, col in enumerate(raw_cols):
            letter = get_column_letter(idx + 1)
            c = str(col).strip()
            if pd.isna(col) or c.lower() in ['nan', '', 'unnamed'] or 'unnamed' in c.lower() or c.lower() == 'none': 
                c = letter
            else: c = f"{letter}_{c}"
            base = c; counter = 1
            while c in new_cols: c = f"{base}_{counter}"; counter += 1
            new_cols.append(c)
        df.columns = new_cols
        return df.dropna(how='all', axis=1).dropna(how='all', axis=0)

# ================= 核心统计算法库 =================
def parse_class_string(val_str):
    val_str = str(val_str).replace(" ", "") 
    ignore = ['0', '0.0', 'nan', 'none', '星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日', '体育', '班会', '国学', '美术', '音乐', '大扫除', '休息', '考试', '学情分析']
    if not val_str or val_str.lower() in ignore or re.search(r'(\d{4})[-/年\.](\d{1,2})[-/月\.](\d{1,2})', val_str) or re.search(r'^第[一二三四五六七八九十0-9\s]+周', val_str):
        return None
        
    count = 1.0
    m_num = re.search(r'(\d+(?:\.\d+)?)$', val_str)
    if m_num:
        if m_num.start() == 0: return None
        count = float(m_num.group(1))
        val_str = val_str[:m_num.start()] 
        
    match = re.match(r'^([\u4e00-\u9fa5a-zA-Z]+?)(高[一二三]|初[一二三]|小[一二三四五六])(.*)$', val_str)
    if match: return {'教师姓名': match.group(1), '课程类别': match.group(2) + match.group(3), '课时数': count}
        
    known_types = ['早自', '正大', '正小', '晚自', '自大', '自小', '辅导', '正课', '早读', '晚修']
    for kt in known_types:
        if val_str.endswith(kt): return {'教师姓名': val_str[:-len(kt)], '课程类别': kt, '课时数': count}
            
    if len(val_str) >= 2: return {'教师姓名': val_str, '课程类别': '常规课', '课时数': count}
    return None

# ================= 侧边栏与核心控制台 =================
st.sidebar.markdown('<div style="text-align:center; padding-bottom:10px;"><h2 style="color:#1e3c72; font-weight:bold;">📁 数据控制台</h2></div>', unsafe_allow_html=True)
uploaded_file = st.sidebar.file_uploader("请拖拽或点击上传 Excel (.xlsm/xlsx)", type=["xlsm", "xlsx"])

if uploaded_file is not None and st.session_state['all_sheets'] is None:
    try:
        with st.spinner('【极速深潜模式启动】：正在原生态读取所有隐藏日期...'):
            raw_sheets = pd.read_excel(uploaded_file, sheet_name=None, engine='openpyxl', header=None)
            clean_sheets = {}
            for sheet_name, df in raw_sheets.items(): clean_sheets[sheet_name] = clean_excel_data(df)
            st.session_state['all_sheets'] = clean_sheets
            st.session_state['current_sheet'] = list(clean_sheets.keys())[0]
            st.sidebar.success("✅ 原生数据解析成功！深潜扫描已就绪。")
    except Exception as e:
        st.error(f"严重错误: {e}")

if st.session_state['all_sheets'] is not None:
    valid_classes = [s for s in st.session_state['all_sheets'].keys() if not any(kw in s for kw in ['总表', '分表', '汇总'])]
    
    st.sidebar.markdown("---")
    st.sidebar.markdown("<div class='sidebar-subtitle'>📍 全局坐标与目标时间</div>", unsafe_allow_html=True)
    col_c1, col_c2 = st.sidebar.columns(2)
    with col_c1: t_period = st.text_input("【节次】", value="L")
    with col_c2: t_time = st.text_input("【时间】", value="M")
    col_c3, col_c4 = st.sidebar.columns(2)
    with col_c3: t_start = st.text_input("【排课起】", value="O")
    with col_c4: t_end = st.text_input("【排课止】", value="U")
    
    g_dates = st.sidebar.date_input("🗓️ 选定目标时间段 (某周/发薪周期)", [])
    
    st.session_state['t_period'] = t_period
    st.session_state['t_time'] = t_time
    st.session_state['t_start'] = t_start
    st.session_state['t_end'] = t_end
    st.session_state['g_dates'] = g_dates

    st.sidebar.markdown("---")
    st.sidebar.markdown("<div class='sidebar-subtitle'>🧑‍🏫 单人课表提取器</div>", unsafe_allow_html=True)
    search_input = st.sidebar.text_input("🔍 查找特定教师：", placeholder="例如：聂俊生")
    schedule_range = st.sidebar.radio("📅 课表周次选择", ["全部周次 (长表)", "仅第一周 (经典单周)"])

    if st.sidebar.button("🔎 提取该教师课表", use_container_width=True):
        if search_input.strip() == "": st.sidebar.warning("请先输入教师姓名！")
        else:
            st.session_state['teacher_mode'] = True
            st.session_state['batch_export_mode'] = False
            st.session_state['global_mode'] = False
            st.session_state['search_teacher'] = search_input.strip()
            st.session_state['schedule_range'] = schedule_range
            
    st.sidebar.markdown("---")
    st.sidebar.markdown("<div class='sidebar-subtitle'>📦 批量导出工具箱</div>", unsafe_allow_html=True)
    
    export_format = st.sidebar.radio("选择打包内容与格式", [
        "🗂️ 班主任核对表 (独立文件打包ZIP，发群专用)",
        "📝 班主任核对表 (全校合一，每班1个Sheet)",
        "🧑‍🏫 教师课表 (单表垂直合并打印)",
        "🧑‍🏫 教师课表 (每个人1个底部Sheet)"
    ])
    
    if st.sidebar.button("🚀 一键提取并打包下发", use_container_width=True):
        if "班主任核对表" in export_format and len(g_dates) < 1:
            st.sidebar.error("⚠️ 生成【班主任核对表】必须在上方【全局坐标与目标时间】中选择你要下发的日期范围（某一周）！")
        else:
            st.session_state['batch_export_mode'] = True
            st.session_state['teacher_mode'] = False
            st.session_state['global_mode'] = False
            st.session_state['export_format'] = export_format
            st.session_state['schedule_range'] = schedule_range

    st.sidebar.markdown("---")
    st.sidebar.markdown("<div class='sidebar-subtitle'>🌐 批量课时发薪统计</div>", unsafe_allow_html=True)
    scope = st.sidebar.radio("📌 薪资统计范围选择", ["所有班级 (全校)", "按年级多选", "自定义勾选班级"])
    
    target_classes = []
    if scope == "所有班级 (全校)": target_classes = valid_classes
    elif scope == "按年级多选":
        grades = st.sidebar.multiselect("挑选年级", ["高一", "高二", "高三"], default=["高三"])
        target_classes = [c for c in valid_classes if any(g in c for g in grades)]
    else:
        target_classes = st.sidebar.multiselect("勾选具体的班级", valid_classes, default=valid_classes[:2])

    if st.sidebar.button("💰 生成课时薪资汇总", use_container_width=True, type="primary"):
        if not target_classes: st.sidebar.error("当前没有选定任何班级！")
        else:
            st.session_state['global_mode'] = True
            st.session_state['teacher_mode'] = False
            st.session_state['batch_export_mode'] = False
            st.session_state['g_targets'] = target_classes
            st.session_state['g_scope'] = scope

# ================= 动态顶部导航 =================
if st.session_state['all_sheets'] is not None:
    all_sheet_names = list(st.session_state['all_sheets'].keys())
    directory_data = {
        "总表 & 汇总": [], "高一年级": [], "高二年级": [], 
        "高三年级": [], "其他表单": []
    }
    for name in all_sheet_names:
        if "总" in name or "分表" in name or "汇总" in name: directory_data["总表 & 汇总"].append(name)
        elif "高一" in name: directory_data["高一年级"].append(name)
        elif "高二" in name: directory_data["高二年级"].append(name)
        elif "高三" in name: directory_data["高三年级"].append(name)
        else: directory_data["其他表单"].append(name)

    st.write("")
    for category, buttons in directory_data.items():
        if not buttons: continue 
        empty_space = 10 - len(buttons) if len(buttons) < 10 else 1
        cols = st.columns([1.2] + [1] * len(buttons) + [empty_space]) 
        with cols[0]: st.markdown(f'<div class="row-title">{category} :</div>', unsafe_allow_html=True)
        for i, btn_name in enumerate(buttons):
            with cols[i+1]:
                if st.button(btn_name, key=f"nav_{btn_name}"):
                    st.session_state['current_sheet'] = btn_name
                    st.session_state['global_mode'] = False 
                    st.session_state['teacher_mode'] = False
                    st.session_state['batch_export_mode'] = False
    st.markdown("<hr style='margin: 15px 0px; border: none; border-top: 1px dashed #cbd5e1;'>", unsafe_allow_html=True)

    # ================= 核心视图分支 =================
    
    # 模式一：【批量导出工具箱：班主任核对表ZIP打包 / 教师全校课表打包】
    if st.session_state.get('batch_export_mode'):
        export_mode_type = st.session_state['export_format']
        
        # 1. 班主任专属核对表提取逻辑
        if "班主任核对表" in export_mode_type:
            f_dates = st.session_state['g_dates']
            f_start = f_dates[0]
            f_end = f_dates[1] if len(f_dates) == 2 else f_dates[0]
            
            zip_title = "🗂️ 班主任自动标红核对表 (ZIP压缩包批量生成)" if "ZIP" in export_mode_type else "📝 班主任自动标红核对表 (全校集合生成)"
            st.markdown(f"<h3 style='color:#1e3c72;'>{zip_title} 📅 ({f_start} 至 {f_end})</h3>", unsafe_allow_html=True)
            
            p_idx = col2num(st.session_state['t_period'])
            t_idx = col2num(st.session_state['t_time'])
            start_idx = col2num(st.session_state['t_start'])
            end_idx = col2num(st.session_state['t_end'])
            
            class_grids = {}
            valid_classes_to_search = [s for s in st.session_state['all_sheets'].keys() if not any(kw in s for kw in ['总表', '分表', '汇总'])]
            
            with st.spinner('正在切片全校各班的标准排课数据，并注入【智能统计与锁定】双重核弹级功能...'):
                for s_name in valid_classes_to_search:
                    s_df = st.session_state['all_sheets'][s_name]
                    if len(s_df.columns) <= max(p_idx, t_idx, end_idx): continue
                    
                    blocks = []; current_block = None; current_date = None
                    
                    for row_idx in range(len(s_df)):
                        row_all_text = "".join([str(x) for x in s_df.iloc[row_idx, :].values])
                        if re.search(r'第[一二三四五六七八九十0-9\s]+周', row_all_text):
                            if current_block is not None and current_date is not None:
                                blocks.append((current_date, current_block))
                            current_block = None
                            current_date = None
                            continue
                            
                        val_str = str(s_df.iloc[row_idx, start_idx]).strip()
                        m = re.search(r'(\d{4})[-/年\.](\d{1,2})[-/月\.](\d{1,2})', val_str)
                        if m:
                            if current_block is not None and current_date is not None: 
                                blocks.append((current_date, current_block))
                            date_str = f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
                            try: current_date = pd.to_datetime(date_str).date()
                            except: current_date = None
                            current_block = [row_idx]
                        else:
                            if current_block is not None: current_block.append(row_idx)
                            
                    if current_block is not None and current_date is not None:
                        blocks.append((current_date, current_block))
                    
                    target_block = None
                    for date, rows in blocks:
                        if date and (f_start <= date <= f_end):
                            target_block = rows; break
                            
                    if target_block:
                        cols_to_extract = [p_idx, t_idx] + list(range(start_idx, end_idx + 1))
                        grid_df = s_df.iloc[target_block, cols_to_extract].copy()
                        grid_df = grid_df.fillna("")
                        
                        cleaned_rows = []
                        for i in range(len(grid_df)):
                            if i == 0 or i == 1:
                                cleaned_rows.append(i) 
                                continue
                                
                            row_vals = [str(x).strip().lower() for x in grid_df.iloc[i, :].values]
                            is_all_empty = all((not v or v in ['0', '0.0', 'nan', 'none', 'nat', 'null']) for v in row_vals)
                            if is_all_empty: continue
                                
                            p_val = row_vals[0]
                            t_val = row_vals[1]
                            if bool(re.search(r'(第.*周|日期|交表|备注|签章|签字|核对)', p_val + t_val)): continue
                                
                            cleaned_rows.append(i)
                                
                        if cleaned_rows:
                            grid_df = grid_df.iloc[cleaned_rows].copy().reset_index(drop=True)
                            
                            for c_idx in range(2, len(grid_df.columns)):
                                val = str(grid_df.iloc[0, c_idx])
                                m = re.search(r'(\d{4})[-/年\.](\d{1,2})[-/月\.](\d{1,2})', val)
                                if m:
                                    grid_df.iloc[0, c_idx] = f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
                                    
                            class_grids[s_name] = grid_df
                        
            if class_grids:
                if "ZIP" in export_mode_type:
                    zip_data = convert_verification_to_zip(class_grids, f_start, f_end)
                    st.success(f"🎉 成功生成！已为您把全校 {len(class_grids)} 个班级的独立表格打包成 ZIP！【已加锁保护、带统计引擎】")
                    st.download_button(
                        label=f"⬇️ 立即下载 ZIP 压缩包 (发群专用)",
                        data=zip_data, 
                        file_name=f"班主任课时核对表大合集_{f_start}至{f_end}.zip",
                        mime="application/zip"
                    )
                    st.info("💡 解压这个 ZIP 文件后，每个 Excel 都内置了【智能统计器】，班主任修改底部的白色格子后，不仅自动变红底黑字，最下面还会帮你算出总共变了多少节！")
                else:
                    excel_data = convert_verification_dfs_to_excel(class_grids, f_start, f_end)
                    st.success(f"🎉 成功生成！全校 {len(class_grids)} 个班级已收录在同一个 Excel 文件的不同 Sheet 里！")
                    st.download_button(
                        label=f"⬇️ 下载全校合并版 Excel",
                        data=excel_data, 
                        file_name=f"班主任课时核对表_全校集合版_{f_start}至{f_end}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.warning(f"😔 在您选定的日期 {f_start} 到 {f_end} 之间，没有找到有效的排课区域。")

        # 2. 教师专属课表提取逻辑 
        else:
            is_first_week_only = "第一周" in st.session_state.get('schedule_range', '')
            title_suffix = "【单周课表】" if is_first_week_only else "【全周期课表】"
            st.markdown(f"<h3 style='color:#1e3c72;'>📦 全校教师专属网格 {title_suffix} 批量打包</h3>", unsafe_allow_html=True)
            
            p_idx = col2num(st.session_state['t_period'])
            t_idx = col2num(st.session_state['t_time'])
            start_idx = col2num(st.session_state['t_start'])
            end_idx = col2num(st.session_state['t_end'])
            
            all_teachers_schedule = []
            valid_classes_to_search = [s for s in st.session_state['all_sheets'].keys() if not any(kw in s for kw in ['总表', '分表', '汇总'])]
            
            with st.spinner('正在像雷达一样扫描全校所有班级数据，为您梳理全校教师信息...'):
                for s_name in valid_classes_to_search:
                    s_df = st.session_state['all_sheets'][s_name]
                    if len(s_df.columns) <= max(p_idx, t_idx, end_idx): continue
                        
                    for col_idx in range(start_idx, end_idx + 1):
                        current_date = None; current_weekday = ""
                        for row_idx in range(len(s_df)):
                            val_str = str(s_df.iloc[row_idx, col_idx]).strip()
                            
                            row_all_text = "".join([str(x) for x in s_df.iloc[row_idx, :].values])
                            if re.search(r'第[一二三四五六七八九十0-9\s]+周', row_all_text):
                                current_date = None
                                continue
                                
                            m = re.search(r'(\d{4})[-/年\.](\d{1,2})[-/月\.](\d{1,2})', val_str)
                            if m:
                                date_str = f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
                                try: current_date = pd.to_datetime(date_str).date()
                                except: pass
                                current_weekday = ""
                                if "星期" in val_str: current_weekday = val_str
                                continue 
                            if "星期" in val_str and len(val_str) <= 10:
                                current_weekday = val_str; continue
                                
                            if not current_date: continue
                            parsed = parse_class_string(val_str)
                            if parsed:
                                teacher_name = parsed['教师姓名']
                                period = str(s_df.iloc[row_idx, p_idx]).strip() if p_idx < len(s_df.columns) else ""
                                time_val = str(s_df.iloc[row_idx, t_idx]).strip() if t_idx < len(s_df.columns) else ""
                                if period.lower() in ['nan', 'none', '0', '0.0']: period = ''
                                if time_val.lower() in ['nan', 'none', '0', '0.0']: time_val = ''
                                time_val = re.sub(r'[-—~]+', '-', time_val)
                                course_name = f"【{s_name}】\n{val_str}" 
                                
                                all_teachers_schedule.append({
                                    '教师姓名': teacher_name, '日期': current_date, '星期': current_weekday,
                                    '节次/分类': period, '时间/参数': time_val, '课程内容': course_name,
                                    '排序辅助': time_val if time_val else "99:99" 
                                })
                                
            if all_teachers_schedule:
                if is_first_week_only: all_teachers_schedule = filter_first_week(all_teachers_schedule)
                ts_df = pd.DataFrame(all_teachers_schedule)
                
                def format_date(row):
                    d_str = row['日期'].strftime('%Y-%m-%d')
                    w_str = row['星期']
                    if not w_str or '星期' not in w_str:
                        weekdays_map = {0: '一', 1: '二', 2: '三', 3: '四', 4: '五', 5: '六', 6: '日'}
                        w_str = f"星期{weekdays_map[row['日期'].weekday()]}"
                    m_wk = re.search(r'(星期[一二三四五六日])', w_str)
                    if m_wk: w_str = m_wk.group(1)
                    return f"{d_str}\n{w_str}"
                    
                ts_df['日期排版'] = ts_df.apply(format_date, axis=1)
                teacher_names = sorted(ts_df['教师姓名'].unique())
                df_dict = {}; df_list = [] 
                progress_bar = st.progress(0); status_text = st.empty()
                
                for i, teacher in enumerate(teacher_names):
                    status_text.text(f"🚀 正在绘制 【{teacher}】 的网格课表 ({i+1}/{len(teacher_names)})...")
                    t_df = ts_df[ts_df['教师姓名'] == teacher]
                    t_group = t_df.groupby(['节次/分类', '时间/参数', '排序辅助', '日期排版'])['课程内容'].apply(lambda x: '\n\n'.join(x.unique())).reset_index()
                    
                    grid_df = pd.pivot_table(t_group, values='课程内容', index=['排序辅助', '节次/分类', '时间/参数'], columns='日期排版', aggfunc=lambda x: '\n\n'.join(x.dropna().unique())).fillna('')
                    grid_df = grid_df.sort_index(level='排序辅助').reset_index(level='排序辅助', drop=True)
                    grid_df.index.names = ['节次', '时间/姓名']; grid_df.columns.name = None
                    
                    title = f"【{teacher}】老师专属课表" + (" (首周)" if is_first_week_only else "")
                    df_dict[teacher] = (grid_df, title); df_list.append((teacher, grid_df, title))
                    progress_bar.progress((i + 1) / len(teacher_names))
                    
                status_text.text("💾 正在按照您要求的格式写入 Excel 文件，马上完成...")
                if "单表" in export_mode_type: excel_data = convert_stacked_dfs_to_excel_pro(df_list, sheet_name="全校打印总表")
                else: excel_data = convert_multiple_dfs_to_excel_pro(df_dict)
                    
                status_text.empty(); progress_bar.empty()
                st.success(f"🎉 终极打包完成！系统已成功梳理出全校 **{len(teacher_names)}** 位教师！空行已全面清退！")
                download_name = f"全校教师课表_打印汇总版_{'首周' if is_first_week_only else '全周期'}.xlsx" if "单表" in export_mode_type else f"全校教师课表_分Sheet装_{'首周' if is_first_week_only else '全周期'}.xlsx"
                
                st.download_button(label=f"⬇️ 立即下载《{download_name}》", data=excel_data, file_name=download_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # 模式二：【教师单人二维网格课表】 
    elif st.session_state['teacher_mode']:
        target_teacher = st.session_state['search_teacher']
        is_first_week_only = "第一周" in st.session_state.get('schedule_range', '')
        title_suffix = "【单周课表】" if is_first_week_only else "【全周期课表】"
        st.markdown(f"<h3 style='color:#1e3c72;'>🧑‍🏫 【{target_teacher}】专属网格 {title_suffix}</h3>", unsafe_allow_html=True)
        
        p_idx = col2num(st.session_state['t_period'])
        t_idx = col2num(st.session_state['t_time'])
        start_idx = col2num(st.session_state['t_start'])
        end_idx = col2num(st.session_state['t_end'])
        
        teacher_schedule = []
        valid_classes_to_search = [s for s in st.session_state['all_sheets'].keys() if not any(kw in s for kw in ['总表', '分表', '汇总'])]
        
        with st.spinner('正在为您提取数据...'):
            for s_name in valid_classes_to_search:
                s_df = st.session_state['all_sheets'][s_name]
                if len(s_df.columns) <= max(p_idx, t_idx, end_idx): continue
                    
                for col_idx in range(start_idx, end_idx + 1):
                    current_date = None
                    current_weekday = ""
                    
                    for row_idx in range(len(s_df)):
                        val_str = str(s_df.iloc[row_idx, col_idx]).strip()
                        
                        row_all_text = "".join([str(x) for x in s_df.iloc[row_idx, :].values])
                        if re.search(r'第[一二三四五六七八九十0-9\s]+周', row_all_text):
                            current_date = None
                            continue
                                
                        m = re.search(r'(\d{4})[-/年\.](\d{1,2})[-/月\.](\d{1,2})', val_str)
                        if m:
                            date_str = f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
                            try: current_date = pd.to_datetime(date_str).date()
                            except: pass
                            current_weekday = ""
                            if "星期" in val_str: current_weekday = val_str
                            continue 
                            
                        if "星期" in val_str and len(val_str) <= 10:
                            current_weekday = val_str
                            continue
                            
                        if not current_date: continue
                        
                        if target_teacher in val_str:
                            period = str(s_df.iloc[row_idx, p_idx]).strip() if p_idx < len(s_df.columns) else ""
                            time_val = str(s_df.iloc[row_idx, t_idx]).strip() if t_idx < len(s_df.columns) else ""
                            if period.lower() in ['nan', 'none', '0', '0.0']: period = ''
                            if time_val.lower() in ['nan', 'none', '0', '0.0']: time_val = ''
                            time_val = re.sub(r'[-—~]+', '-', time_val)
                            course_name = f"【{s_name}】\n{val_str}" 
                            
                            teacher_schedule.append({
                                '日期': current_date,
                                '星期': current_weekday,
                                '节次/分类': period,
                                '时间/参数': time_val,
                                '课程内容': course_name,
                                '排序辅助': time_val if time_val else "99:99" 
                            })
                            
        if teacher_schedule:
            if is_first_week_only: teacher_schedule = filter_first_week(teacher_schedule)

            ts_df = pd.DataFrame(teacher_schedule)
            
            def format_date(row):
                d_str = row['日期'].strftime('%Y-%m-%d')
                w_str = row['星期']
                if not w_str or '星期' not in w_str:
                    weekdays_map = {0: '一', 1: '二', 2: '三', 3: '四', 4: '五', 5: '六', 6: '日'}
                    w_str = f"星期{weekdays_map[row['日期'].weekday()]}"
                m_wk = re.search(r'(星期[一二三四五六日])', w_str)
                if m_wk: w_str = m_wk.group(1)
                return f"{d_str}\n{w_str}"
                
            ts_df['日期排版'] = ts_df.apply(format_date, axis=1)
            ts_df = ts_df.groupby(['节次/分类', '时间/参数', '排序辅助', '日期排版'])['课程内容'].apply(lambda x: '\n\n'.join(x.unique())).reset_index()
            
            grid_df = pd.pivot_table(
                ts_df, values='课程内容', index=['排序辅助', '节次/分类', '时间/参数'], 
                columns='日期排版', aggfunc=lambda x: '\n\n'.join(x.dropna().unique())
            ).fillna('')
            
            grid_df = grid_df.sort_index(level='排序辅助').reset_index(level='排序辅助', drop=True)
            grid_df.index.names = ['节次', '时间/姓名'] 
            grid_df.columns.name = None
            
            st.success(f"🎉 找齐了！共为您提取出【{target_teacher}】的专属课表：")
            st.dataframe(grid_df, use_container_width=True)
            
            formal_title = f"【{target_teacher}】专属课表" + (" (首周)" if is_first_week_only else " (全周期)")
            excel_data = convert_df_to_excel_pro(grid_df, sheet_name="个人课表", title=formal_title)
            st.download_button(
                label=f"⬇️ 下载《{target_teacher}专属课表》",
                data=excel_data, file_name=f"{target_teacher}_{'首周' if is_first_week_only else '全周期'}课表.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning(f"😔 没有找到包含【{target_teacher}】的排课信息！")

    # 模式三：【全局发薪汇总统计】 
    elif st.session_state['global_mode']:
        f_dates = st.session_state['g_dates']
        targets = st.session_state['g_targets']
        start_idx = col2num(st.session_state['t_start'])
        end_idx = col2num(st.session_state['t_end'])
        
        report_title_prefix = "全校" if st.session_state['g_scope'] == "所有班级 (全校)" else "选中班级"
        sub_title = f"📅 ({f_dates[0]} 至 {f_dates[1]})" if len(f_dates)==2 else (f"📅 ({f_dates[0]})" if len(f_dates)==1 else "📅 (全学期汇总)")
        st.markdown(f"<h3 style='color:#1e3c72;'>🌐 【{report_title_prefix}】课时总汇 {sub_title}</h3>", unsafe_allow_html=True)
        
        all_records = []
        with st.spinner('正在执行全盘数据拉取...'):
            for s_name in targets:
                if s_name not in st.session_state['all_sheets']: continue
                s_df = st.session_state['all_sheets'][s_name]
                if len(s_df.columns) <= max(start_idx, end_idx): continue
                
                for col_idx in range(start_idx, end_idx + 1):
                    current_date = None
                    for row_idx in range(len(s_df)):
                        val_str = str(s_df.iloc[row_idx, col_idx]).strip()
                        
                        row_all_text = "".join([str(x) for x in s_df.iloc[row_idx, :].values])
                        if re.search(r'第[一二三四五六七八九十0-9\s]+周', row_all_text):
                            current_date = None
                            continue
                            
                        m = re.search(r'(\d{4})[-/年\.](\d{1,2})[-/月\.](\d{1,2})', val_str)
                        if m:
                            date_str = f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
                            try: current_date = pd.to_datetime(date_str).date()
                            except: pass
                            continue
                        
                        if current_date:
                            if len(f_dates) == 2:
                                if not (f_dates[0] <= current_date <= f_dates[1]): continue
                            elif len(f_dates) == 1:
                                if current_date != f_dates[0]: continue
                                
                            parsed = parse_class_string(val_str)
                            if parsed:
                                parsed['来源班级'] = s_name
                                parsed['来源日期'] = str(current_date)
                                all_records.append(parsed)
                            
        if all_records:
            stat_df = pd.DataFrame(all_records)
            pivot_df = pd.pivot_table(stat_df, values='课时数', index='教师姓名', columns='课程类别', aggfunc='sum', fill_value=0)
            pivot_df['总计'] = pivot_df.sum(axis=1)
            
            st.success(f"🎉 统计完毕！共 {len(stat_df['教师姓名'].unique())} 位老师，总计 {stat_df['课时数'].sum()} 节。")
            st.dataframe(pivot_df, use_container_width=True)
            
            formal_title = f"【{report_title_prefix}汇总】课时报表 {sub_title.replace('📅 ', '')}"
            excel_data = convert_df_to_excel_pro(pivot_df, sheet_name="数据汇总", title=formal_title)
            st.download_button(
                label=f"⬇️ 导出《{report_title_prefix}汇报表格》为 Excel",
                data=excel_data, file_name=f"{report_title_prefix}课时报表.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("⚠️ 在指定的范围中，未抓取到有效课时！")
            
    # 模式四：【单班级管理视图】
    else:
        current = st.session_state['current_sheet']
        st.markdown(f"<h4 style='color:#1e3c72;'>👁️ 当前查看 : 【 {current} 】</h4>", unsafe_allow_html=True)
        
        df_current = st.session_state['all_sheets'][current].copy()
        display_df = df_current.astype(str).replace({' 00:00:00': ''}, regex=True).replace({'nan': '', 'None': ''})
        st.dataframe(display_df, use_container_width=True, height=350)

        st.markdown("---")
        tab1, tab2 = st.tabs(["📏 【排课表】按列字母锁定统计", "📊 【明细表】手动选列统计"])
        
        with tab1:
            start_idx = col2num(st.session_state['t_start'])
            end_idx = col2num(st.session_state['t_end'])
            
            if start_idx <= end_idx and len(display_df.columns) > end_idx:
                all_dates_in_range = set()
                for col_idx in range(start_idx, end_idx + 1):
                    for row_idx in range(len(display_df)):
                        val = display_df.iloc[row_idx, col_idx]
                        m = re.search(r'(\d{4})[-/年\.](\d{1,2})[-/月\.](\d{1,2})', str(val).strip())
                        if m:
                            date_str = f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
                            try: all_dates_in_range.add(pd.to_datetime(date_str).date())
                            except: pass
                
                if all_dates_in_range:
                    min_d, max_d = min(all_dates_in_range), max(all_dates_in_range)
                    date_range = st.date_input(f"🗓️ 选择提取区间 (默认全选)：", [min_d, max_d])
                    
                    if st.button("🚀 开始本班精准提取", type="primary"):
                        if len(date_range) >= 1:
                            f_start = date_range[0]
                            f_end = date_range[1] if len(date_range) == 2 else date_range[0]
                            
                            records = []
                            for col_idx in range(start_idx, end_idx + 1):
                                current_date = None
                                for row_idx in range(len(display_df)):
                                    val_str = str(display_df.iloc[row_idx, col_idx]).strip()
                                    
                                    row_all_text = "".join([str(x) for x in display_df.iloc[row_idx, :].values])
                                    if re.search(r'第[一二三四五六七八九十0-9\s]+周', row_all_text):
                                        current_date = None
                                        continue
                                        
                                    m = re.search(r'(\d{4})[-/年\.](\d{1,2})[-/月\.](\d{1,2})', val_str)
                                    if m:
                                        date_str = f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
                                        try: current_date = pd.to_datetime(date_str).date()
                                        except: pass
                                        continue
                                    
                                    if current_date and (f_start <= current_date <= f_end):
                                        parsed = parse_class_string(val_str)
                                        if parsed: records.append(parsed)
                                            
                            if records:
                                stat_df = pd.DataFrame(records)
                                pivot_df = pd.pivot_table(stat_df, values='课时数', index='教师姓名', columns='课程类别', aggfunc='sum', fill_value=0)
                                pivot_df['总计'] = pivot_df.sum(axis=1)
                                
                                st.success(f"🎉 统计完毕！【{current}】共计 {stat_df['课时数'].sum()} 节课时。")
                                st.dataframe(pivot_df, use_container_width=True)
                                
                                formal_title = f"【{current}】课时统计报表 ({f_start}至{f_end})"
                                excel_data = convert_df_to_excel_pro(pivot_df, sheet_name=current, title=formal_title)
                                st.download_button(
                                    label=f"⬇️ 导出《{current}报表》",
                                    data=excel_data, file_name=f"{current}_课时报表.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                            else:
                                st.warning("未找到可识别的课时。")
                else:
                    st.warning("⚠️ 没有在指定的列范围内扫描到包含日期的行！")
            else:
                st.warning("您侧边栏填写的字母超出了表格范围，请检查！")

        with tab2:
            available_cols = list(display_df.columns)
            def guess_index(kw):
                for i, c in enumerate(available_cols):
                    if any(k in str(c) for k in kw): return i
                return 0
                
            col1, col2, col3 = st.columns(3)
            with col1: name_col = st.selectbox("👤 【姓名】列", available_cols, index=guess_index(['姓名','教师']))
            with col2: type_col = st.selectbox("🏷️ 【类别】列", available_cols, index=guess_index(['子类','类别']))
            with col3: count_col = st.selectbox("🔢 【数量】列", available_cols, index=guess_index(['课数','课时']))
                
            if st.button("📊 生成常规统计"):
                try:
                    stat_df = df_current.copy()
                    stat_df[count_col] = pd.to_numeric(stat_df[count_col], errors='coerce').fillna(0)
                    stat_df = stat_df[stat_df[name_col].notna()]
                    stat_df = stat_df[stat_df[name_col].astype(str).str.strip() != '']
                    pivot_df = pd.pivot_table(stat_df, values=count_col, index=name_col, columns=type_col, aggfunc='sum', fill_value=0)
                    pivot_df['总计'] = pivot_df.sum(axis=1)
                    st.dataframe(pivot_df, use_container_width=True)
                    
                    formal_title = f"【{current}】常规课时统计"
                    excel_data = convert_df_to_excel_pro(pivot_df, sheet_name=current, title=formal_title)
                    st.download_button(
                        label="⬇️ 导出常规报表", data=excel_data, file_name=f"{current}_常规课时.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except:
                    st.warning("无法生成，请确认选对了列名！")
else:
    st.info("👆 请先在左侧上传您的 Excel 文件！")
